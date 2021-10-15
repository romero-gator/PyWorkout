from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
from openpyxl.workbook import workbook

def handleWorkout(startRow, endRow, startCol, endCol, sheet):
    for col in sheet.iter_cols(min_row=startRow, max_row=endRow, min_col=startCol, max_col=endCol):
        sesh = 0
        for cell in col:
            if isinstance(cell.value, str):
                cellString = str(cell.value)
                if cellString.rfind('@') != -1:
                    power = 0
                    split0 = cellString.partition(', ')

                    while all(split0):
                        split1 = split0[0].partition('x')
                        split2 = split1[2].partition('@')
                        # make a check statement to ensure the splits can be converted to int
                        power = power + (int(split1[0]) * int(split2[0]) * int(split2[2]))
                        split0 = split0[2].partition(', ')
                    
                    split1 = split0[0].partition('x')
                    split2 = split1[2].partition('@')
                    # make a check statement to ensure the splits can be converted to int
                    power = power + (int(split1[0]) * int(split2[0]) * int(split2[2]))
                    sesh = sesh + power

                    workout = str(sheet.cell(cell.row, 1).value)
                    if workout.rfind('(') == -1:
                        sheet.cell(cell.row, 1).value = workout + " (" + str(power) + ")"
                    else:
                        first = workout.partition(' (')
                        second = first[2].partition(')')
                        max = int(second[0])
                        if max < power:
                            sheet.cell(cell.row, 1).value = first[0] + " (" + str(power) +")"
        if col[0].column != 1:
            workoutName = sheet.cell(startRow, 1).value
            workoutName = workoutName.partition(' (')
            workoutNum = workoutName[2].partition(')')
            bestSesh = int(workoutNum[0])
            if sesh > bestSesh:
                bestSesh = sesh
                sheet.cell(startRow, 1).value = workoutName[0] + " (" + str(sesh) + ")"
            # change color in the top of sesh box
            colorNum = sesh / bestSesh
            sheet.cell(startRow, cell.column).value = round((colorNum * 100), 1)
            if colorNum == 0.0:
                colorNum = "abb2b9"
            elif colorNum > 0.5:
                colorNum = hex(int((1 - (colorNum / 1.5)) * 255))
                colorNum = colorNum.partition('x')
                colorNum = colorNum[2] + "ff00"
            elif colorNum < 0.5:
                colorNum = hex(int((colorNum * 2) * 255))
                colorNum = colorNum.partition('x')
                colorNum = "ff" + colorNum[2] + "00"
            else:
                colorNum = "ffff00"
            sheet.cell(startRow, cell.column).fill = PatternFill("solid", fgColor=colorNum)

wb = load_workbook('workouts.xlsx')
sheet = wb.active


# 00FF00 green          1.0
# xxFF00                x           (1 - (x / 1.5)) * 255
# FFFF00 yellow         0.5
# FFxx00                x           (x * 2) * 255
# FF0000 red            0.0

chestRow = 0
tricepsRow = 0
shouldersRow = 0
backRow = 0
bicepsRow = 0
legsRow = 0

for row in sheet.iter_rows(min_row=1, max_row=59, min_col=1, max_col=1):
    if "CHEST" in row[0].value:
        chestRow = row[0].row
    elif "TRICEPS" in row[0].value:
        tricepsRow = row[0].row
    elif "SHOULDERS" in row[0].value:
        shouldersRow = row[0].row
    elif "BACK" in row[0].value:
        backRow = row[0].row
    elif "BICEPS" in row[0].value:
        bicepsRow = row[0].row
    elif "LEGS" in row[0].value:
        legsRow = row[0].row

# BACK
handleWorkout(28,39,1,4,sheet)


wb.save('workouts.xlsx')